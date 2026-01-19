from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font

def format_excel(path: str):
    wb = load_workbook(path)
    ws = wb.active

    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="thick", color="000000")

    # Centralizar tudo + borda fina
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Bordas externas grossas
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            border = Border(
                top=thick if r == 1 else thin,
                left=thick if c == 1 else thin,
                right=thick if c == max_col else thin,
                bottom=thick if r == max_row else thin,
            )
            cell.border = border

    # Cabeçalho vermelho e texto branco
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Ajustar largura automática das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # A, B, C...
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        # Multiplicador para dar um pouco de espaço extra
        adjusted_width = (max_length + 5)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(path)
